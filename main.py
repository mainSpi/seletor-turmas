import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import random
import time

seed = time.time()
random.seed(seed)
print("SEED: " + str(seed))

lugares = {
    "Heitor Beltrão": 8,
    "João Barros Neto" : 8,
    "Hesfa/Marcolino" : 10,
    "Estácio de Sá" : 8,
    "Felippe Cardoso" : 14
}

turmas = ["t1.csv", "t2.csv"]


turmas = {list(turmas)[i]: [] for i in range(len(turmas))} # cria um array vazio para cada turma
stats = {list(turmas)[i]: {j:0 for j in range(1 , len(lugares.keys())+1)} for i in range(len(turmas))}

def salvarComoPlanilha():
    wb = Workbook()
    wb.remove(wb.active)
    for turma, divisao in turmas.items():
        ws = wb.create_sheet(turma)

        for lugar, alunos in divisao.items():
            coluna = list(divisao.keys()).index(lugar) + 1
            celulaLugar = ws.cell(row=1, column=coluna)
            celulaLugar.value = lugar
            celulaLugar.font = Font(color = "FF0000")

            for i in range(0, len(alunos)):
                ws.cell(row=i+2, column=coluna).value = alunos[i]

        # https://www.jquery-az.com/auto-adjust-column-width-by-python-openpyxl/
        for col in ws.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > length:
                    length = len(str(cell.value))
            ws.column_dimensions[column].width = length - 3 # ele exagera um pouco seila

    ws = wb.create_sheet("stats")

    ws.cell(row=1, column=1).value = "Prioridades/Repetições"
    for i in range(1 , len(lugares.keys())+1):
        ws.cell(row=i+1, column=1).value = i

    for turma, ps in stats.items():
        coluna = list(stats.keys()).index(turma) + 2
        ws.cell(row=1, column=coluna).value = turma

        valores = list(ps.values())
        for i in range(0, len(valores)):
            ws.cell(row=i+2, column=coluna).value = valores[i]

    wb.save("saida.xlsx")

for turma in turmas.keys():
    divisao = {list(lugares.keys())[i]: [] for i in range(len(lugares.keys()))} # # cria um array vazio para cada lugar
    
    with open(turma) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        alunos = []

        for row in csv_reader:
            alunos.append(row)

        random.shuffle(alunos)

        for aluno in alunos:
            prioridades = aluno[1:len(lugares.keys())+1] # vem umas colunas em branco que podem atrapalhar
            prioridades = list(map(int, prioridades)) # transformando em numero
            
            for i in range(1, len(lugares.keys()) + 1): # loopando sobre as prioridades ate achar
                index = prioridades.index(i)
                lugar = list(lugares.keys())[index]
                temVaga = True if len(divisao[lugar]) < lugares[lugar] else False
                if (temVaga):
                    divisao[lugar].append(aluno[0])
                    stats[turma][i] += 1
                    break # achou vaga, sai do loop

    turmas[turma] = divisao

salvarComoPlanilha()